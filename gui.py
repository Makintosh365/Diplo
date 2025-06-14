import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
# core-модули
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "core")))
from core.preprocessing import extract_features, load_csv_files, save_processed_data
from core.labeling import add_labels
from core.model import BotnetClassifier


class TrafficApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Анализ трафика с помощью нейронной сети")
        self.geometry("700x500")
        self.configure(bg="#1c1f2b")

        self.selected_file = None
        self.result_df = None

        self.create_widgets()

    def create_widgets(self):
        title = tk.Label(self, text="Трафик анализатор", font=("Arial", 24, "bold"), fg="white", bg="#1c1f2b")
        subtitle = tk.Label(self, text="", font=("Arial", 12), fg="lightgray", bg="#1c1f2b")
        title.pack(pady=(20, 0))
        subtitle.pack(pady=(0, 20))
        frame = tk.Frame(self, bg="#1c1f2b")
        frame.pack()
        # === ANALYZE TRAFFIC ===
        analyze = tk.LabelFrame(frame, text="Анализ трафика", bg="#2a2e3f", fg="white", font=("Arial", 10, "bold"), padx=20, pady=20)
        analyze.grid(row=0, column=0, padx=20, pady=10)
        self.label_path = tk.Label(analyze, text="Файл не выбран", fg="white", bg="#2a2e3f")
        self.label_path.pack()
        ttk.Button(analyze, text="Загрузить CSV-файл", command=self.load_file).pack(pady=5)
        ttk.Button(analyze, text="Сгенерировать отчет", command=self.generate_report).pack()
        # === TRAIN NETWORK ===
        train = tk.LabelFrame(frame, text="Обучение нейроной сети", bg="#2a2e3f", fg="white", font=("Arial", 10, "bold"), padx=20, pady=20)
        train.grid(row=0, column=1, padx=20, pady=10)
        ttk.Button(train, text="Начать обучение", command=self.retrain_model).pack(pady=30)
        # === SAVE REPORT ===
        ttk.Button(self, text="Сохранить отчет", command=self.save_result).pack(pady=10)
        # === OUTPUT ===
        self.text = tk.Text(self, height=12, bg="#14151c", fg="white", insertbackground="white")
        self.text.pack(fill="both", padx=20, pady=10)
        
    def log(self, msg: str):
        self.text.insert("end", msg + "\n")
        self.text.see("end")

    def load_file(self):
        path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if path:
            self.selected_file = path
            self.label_path.config(text=os.path.basename(path))
            self.log(f"📥 Загружен файл: {path}")


    def generate_report(self):
        if not self.selected_file:
            messagebox.showwarning("Нет файла", "Сначала выберите CSV-файл.")
            return
        try:
            chunk_iter = pd.read_csv(
                self.selected_file,
                low_memory=False,
                chunksize=50000,
            )

            model = BotnetClassifier()
            vectorizer = None
            scaler = None
            result_chunks = []

            for chunk in chunk_iter:
                chunk["source_file"] = os.path.basename(self.selected_file)
                df_feat, vectorizer, scaler = extract_features(
                    chunk.copy(), vectorizer, scaler
                )

                if model.model is None:
                    model.load_model(input_size=df_feat.shape[1] - 1)

                preds = model.predict(df_feat)
                chunk["predicted_label"] = preds
                result_chunks.append(chunk)

            df_raw = pd.concat(result_chunks, ignore_index=True)
            self.result_df = df_raw

            # Подозрительные строки
            suspicious = df_raw[df_raw["predicted_label"] != "normal"]
            if suspicious.empty:
                self.log("🟢 Подозрительных пакетов не найдено.")
                return

            # 🔽 Сохраняем вручную
            out_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Сохранить отчёт как..."
            )
            if not out_path:
                self.log("🚫 Сохранение отменено.")
                return

            # 🎨 Цвета по метке
            label_colors = {
                "botnet_http":     "ADD8E6",  # light blue
                "botnet_icmp":     "90EE90",  # light green
                "botnet_tcp":      "FFD580",  # light orange
                "botnet_slowloic": "FF9999",  # light red
            }

            # 🧾 Генерация Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Suspicious Packets"

            for r in dataframe_to_rows(suspicious, index=False, header=True):
                ws.append(r)

            ws.auto_filter.ref = ws.dimensions

            label_col_idx = None
            for i, cell in enumerate(ws[1], start=1):
                if cell.value == "predicted_label":
                    label_col_idx = i
                    break

            # Подсветка строк
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                label = row[label_col_idx - 1].value
                color = label_colors.get(label, "DDDDDD")  # default gray
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for cell in row:
                    cell.fill = fill

            wb.save(out_path)
            self.log(f"✅ Отчёт сохранён: {out_path}")

            summary = df_raw["predicted_label"].value_counts()
            self.log("📊 Сводка по меткам:")
            self.log(summary.to_string())

        except Exception as e:
            self.log(f"❌ Ошибка: {str(e)}")



    def retrain_model(self):
        try:
            self.log("📥 Загружаем данные из data/raw/")
            df = load_csv_files("data/raw")
            df, _, _ = extract_features(df)
            df = add_labels(df)
            save_processed_data(df, "data/processed/train.csv")

            self.log("🧠 Обучаем модель...")
            model = BotnetClassifier()
            model.train(df)
            model.save_model()
            self.log("✅ Модель обучена и сохранена.")
        except Exception as e:
            self.log(f"❌ Ошибка при обучении: {str(e)}")

    def save_result(self):
        if self.result_df is None:
            messagebox.showinfo("Нет результата", "Сначала выполните анализ.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".csv")
        if path:
            self.result_df.to_csv(path, index=False)
            self.log(f"💾 Результат сохранён: {path}")


if __name__ == "__main__":
    app = TrafficApp()
    app.mainloop()
