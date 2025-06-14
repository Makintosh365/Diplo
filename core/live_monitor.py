import os
import sys
import threading
import time
from datetime import datetime

import pandas as pd
from scapy.all import sniff, Packet

import tkinter as tk
from scapy.layers.inet import IP, TCP, UDP, ICMP

# core-модули
from core.preprocessing import extract_features
from core.model import BotnetClassifier

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill


class LiveBotnetMonitor:
    def __init__(self):
    

        self.packet_list = []

        # 📥 Загружаем train.csv, чтобы определить точное число признаков
        try:
            df_train = pd.read_csv("data/processed/train.csv")
            input_size = df_train.drop(columns=["label", "source_file"], errors="ignore").shape[1]
        except Exception as e:
            print(f"❌ Не удалось загрузить train.csv: {e}")
            input_size = 110  # fallback, если вдруг нет файла (можно подставить вручную)

        # 🧠 Загрузка модели
        self.model = BotnetClassifier()
        self.model.load_model(input_size=input_size)

        self.running = True

        # 🖥️ GUI-индикатор
        self.root = tk.Tk()
        self.root.title("Botnet Monitor")
        self.root.geometry("100x100")
        self.root.resizable(False, False)
        self.status_canvas = tk.Canvas(self.root, width=100, height=100, highlightthickness=0)
        self.status_canvas.pack()
        self.status_indicator = self.status_canvas.create_rectangle(10, 10, 90, 90, fill="grey")


    def start_gui_loop(self):
        self.update_indicator("grey")
        threading.Thread(target=self.start_sniffing_loop, daemon=True).start()
        self.root.mainloop()

    def update_indicator(self, color: str):
        self.status_canvas.itemconfig(self.status_indicator, fill=color)

    def start_sniffing_loop(self):
        self.start_time = time.time()

        while self.running:
            self.packet_list = []

            print("📡 Сбор трафика начат (60 секунд)...")
            sniff(timeout=60, prn=self.handle_packet, store=False)

            print(f"📦 Сбор завершён: {len(self.packet_list)} пакетов.")
            self.analyze_packets()

    def handle_packet(self, pkt: Packet):
        try:
            info = str(pkt.summary())
            proto = pkt.proto if hasattr(pkt, "proto") else 0
            length = len(pkt)
            timestamp = pkt.time
            self.packet_list.append({
                "No.": len(self.packet_list) + 1,
                "Time": timestamp,
                "Protocol": proto,
                "Length": length,
                "Info": info,
                "source_file": "live_capture"
            })
        except Exception as e:
            print(f"Ошибка в обработке пакета: {e}")

    def handle_packet(self, pkt: Packet):
        try:
            # Проверяем, что есть IP уровень
            if IP in pkt:
                ip_layer = pkt[IP]
                src = ip_layer.src
                dst = ip_layer.dst
            else:
                src = "?"
                dst = "?"

            # Протокол
            if TCP in pkt:
                proto = "TCP"
            elif UDP in pkt:
                proto = "UDP"
            elif ICMP in pkt:
                proto = "ICMP"
            else:
                proto = pkt.name  # fallback

            # Info из summary
            info = pkt.summary()
            length = len(pkt)
            timestamp = pkt.time
            number = len(self.packet_list) + 1

            self.packet_list.append({
                "No.": number,
                "Time": round(timestamp - self.start_time, 6),
                "Source": src,
                "Destination": dst,
                "Protocol": proto,
                "Length": length,
                "Info": info,
                "source_file": "live_capture"
            })
        except Exception as e:
            print(f"Ошибка обработки пакета: {e}")
    def analyze_packets(self):
        if not self.packet_list:
            return

        df_raw = pd.DataFrame(self.packet_list)
        try:
            df_feat = extract_features(df_raw.copy())
            preds = self.model.predict(df_feat)
            df_raw["predicted_label"] = preds

            botnet_count = (df_raw["predicted_label"] != "normal").sum()
            total = len(df_raw)

            if botnet_count >= total / 3:
                self.update_indicator("red")
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
                filename = f"report_{timestamp}.xlsx"
                self.save_excel_report(df_raw[df_raw["predicted_label"] != "normal"], filename)
                print(f"⚠️ Атака обнаружена! Сохранён отчёт: {filename}")
            else:
                self.update_indicator("green")
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
                filename = f"report_{timestamp}.xlsx"
                self.save_excel_report(df_raw[df_raw["predicted_label"] != "normal"], filename)
                print("✅ Трафик чист.")
        except Exception as e:
            print(f"❌ Ошибка анализа: {e}")

    def save_excel_report(self, df: pd.DataFrame, filename: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Suspicious Packets"

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        ws.auto_filter.ref = ws.dimensions
        label_col = None
        for i, cell in enumerate(ws[1], start=1):
            if cell.value == "predicted_label":
                label_col = i
                break

        label_colors = {
            "botnet_http": "ADD8E6",
            "botnet_icmp": "90EE90",
            "botnet_tcp": "FFD580",
            "botnet_slowloic": "FF9999",
        }

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            label = row[label_col - 1].value
            color = label_colors.get(label, "DDDDDD")
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for cell in row:
                cell.fill = fill

        os.makedirs("data/reports", exist_ok=True)
        wb.save(os.path.join("data/reports", filename))


if __name__ == "__main__":
    app = LiveBotnetMonitor()
    app.start_gui_loop()
